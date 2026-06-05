from __future__ import annotations

import re
import zipfile
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path("/Users/pessoal/Documents/priorizacao-implantacao-dashboard")
OUTPUT = ROOT / "outputs" / "batidas-qualitativas-semana-2026-05-11.xlsx"
WEEK_LABEL = "Semana de 2026-05-11"

DOCS = [
    {
        "implanter": "Aline Andrade",
        "meeting_date": "2026-05-11",
        "path": Path("/Users/pessoal/Downloads/Batida de Funil - Aline A - 2026_05_11 09_11 GMT-03_00 - Anotações do Gemini.docx"),
    },
    {
        "implanter": "David Valoski",
        "meeting_date": "2026-05-11",
        "path": Path("/Users/pessoal/Downloads/Batida de Funil - David - 2026_05_11 15_05 GMT-03_00 - Anotações do Gemini.docx"),
    },
    {
        "implanter": "Maria Elena",
        "meeting_date": "2026-05-11",
        "path": Path("/Users/pessoal/Downloads/Batida de funil - Maria - 2026_05_11 14_15 GMT-03_00 - Anotações do Gemini.docx"),
    },
    {
        "implanter": "Natieli Ambrosi",
        "meeting_date": "2026-05-11",
        "path": Path("/Users/pessoal/Downloads/Batida de Funil - Natieli - 2026_05_11 15_47 GMT-03_00 - Anotações do Gemini.docx"),
    },
    {
        "implanter": "Aline Santos",
        "meeting_date": "2026-05-11",
        "path": Path("/Users/pessoal/Downloads/Batida de Funil -Aline - 2026_05_11 13_15 GMT-03_00 - Anotações do Gemini.docx"),
    },
]

QUALITATIVE_ROWS = [
    ("Aline Andrade", "Nomos", "Atenção", "Projeto travado", "Cliente grande em go live, mas grade segue travada e depende de retorno do time do cliente.", "Aline Andrade", "Hoje", "Pressionar resposta; se não houver retorno até o fim do dia, escalar com Ana Cassia.", "Aberto"),
    ("Aline Andrade", "Maria Margarida", "Saudável", "Upgrade", "Go live iminente e percepção positiva do funil, com chance de subir de plano após uso inicial.", "Aline Andrade", "1-2 semanas", "Acompanhar go live e mapear potencial de upgrade depois do primeiro mês.", "Aberto"),
    ("Aline Andrade", "Ecologic", "Risco", "Contraction", "Cliente deve reduzir usuários e pode ficar perdido temporariamente durante a implantação do OMIE.", "Aline Andrade", "Amanhã", "Discutir downgrade para 10 usuários e registrar eventual perda temporária.", "Aberto"),
    ("Aline Andrade", "Maxil", "Perdido", "Perdido", "Projeto sem engajamento após múltiplas tentativas; operação decidiu seguir protocolo de perdido.", "Aline Andrade", "Esta semana", "Enviar e-mail final formalizando pausa e documentar motivo com clareza.", "Aberto"),
    ("Aline Andrade", "Tecnoluva", "Atenção", "Projeto travado", "Projeto lento por auditoria e muitas regras, mas sem sinal claro de cancelamento.", "Aline Andrade", "Monitoramento", "Seguir em acompanhamento sem tratar como churn no momento.", "Aberto"),
    ("Aline Andrade", "Lecran", "Atenção", "Sem movimento", "Cliente depende de melhoria de roadmap; há desconforto com prazo e necessidade de cobrança interna.", "Aline Andrade", "Quarta-feira", "Cobrar Ed e verificar possibilidade de priorização paga com PAT.", "Aberto"),
    ("Aline Andrade", "Promissão Alimentos", "Atenção", "Projeto travado", "Projeto parado do lado do cliente, mas com evolução interna deles em andamento.", "Aline Andrade", "Esta semana", "Realizar novo follow-up para verificar liberação da implantação.", "Aberto"),
    ("Aline Andrade", "Grupo Pascoal Alimentos", "Atenção", "Ressurrect", "Nova pessoa assumiu a conta e demonstrou boa receptividade ao sistema.", "Aline Andrade", "Esta semana", "Agendar nova conversa e reforçar material de apoio para retomada.", "Aberto"),
    ("Aline Andrade", "Nona Mineira", "Saudável", "Ressurrect", "Projeto foi resgatado de perdido e cliente voltou a usar a ferramenta.", "Aline Andrade", "Monitoramento", "Garantir que o cliente reapareça na base e manter acompanhamento leve.", "Aberto"),
    ("Aline Andrade", "Hábito", "Oportunidade", "Upgrade", "Cliente resolveu entraves e aceitou proposta para mais usuários.", "Aline Andrade", "Após aprovação da diretoria", "Avançar com upgrade de 10 usuários quando houver aprovação final.", "Aberto"),

    ("David Valoski", "Infine Distribuidora", "Atenção", "Conclusão", "Cliente com várias tentativas de agendamento; operação quer concluir por mensagem e nota.", "David Valoski", "Esta semana", "Enviar mensagem final e solicitar nota de conclusão.", "Aberto"),
    ("David Valoski", "Origem Distribuidora", "Saudável", "Conclusão", "Cliente já usava no kickoff e deve ser concluído assim que fechar a janela mínima do projeto.", "David Valoski", "Esta semana", "Validar conclusão e encerrar o projeto.", "Aberto"),
    ("David Valoski", "RFX", "Atenção", "Conclusão", "Cliente já usa, mas segue sem agendar a reunião de encerramento.", "David Valoski", "Esta semana", "Cobrar agendamento para checkpoint de conclusão.", "Aberto"),
    ("David Valoski", "G2 Healthcare", "Atenção", "Sem movimento", "Conta usa a ferramenta, mas conclusão depende de treinamento de novas contratações.", "David Valoski", "Após treinamento", "Não encerrar antes de realizar a reunião pendente.", "Aberto"),
    ("David Valoski", "Bioflorata", "Risco", "Projeto travado", "Integração com Web Mais apresenta inconsistências em produto, estoque e tabela de preços.", "David Valoski", "Urgente", "Escalar parceiro de ERP e manter Ana Cassia no grupo de suporte.", "Aberto"),
    ("David Valoski", "Amendo Pro", "Atenção", "Expansão", "Projeto avança com Tinepay e precisa validar pedido de teste antes do treinamento.", "David Valoski", "Em breve", "Validar configuração financeira e puxar treinamento com vendedores.", "Aberto"),
    ("David Valoski", "Meta Ferragens", "Oportunidade", "Upgrade", "Cliente já fez upgrade de usuários e deve avançar reunião estratégica.", "David Valoski", "Amanhã", "Aproveitar reunião para consolidar expansão e encerrar implantação.", "Aberto"),
    ("David Valoski", "Ful", "Saudável", "Conclusão", "Cliente vinha apontando erros, mas as principais pendências foram resolvidas.", "David Valoski", "Esta semana", "Encerrar o projeto formalmente.", "Aberto"),
    ("David Valoski", "Viva Bela Comercial", "Risco", "Projeto travado", "Projeto depende da correção do mesmo erro de integração da Web Mais para concluir.", "David Valoski", "Urgente", "Concluir assim que o ponto crítico de integração for resolvido.", "Aberto"),
    ("David Valoski", "Corporate Hub", "Saudável", "Conclusão", "Cliente já trocou ERP, está faturando pedidos e ficou muito próximo do encerramento.", "David Valoski", "Esta semana", "Validar andamento final e concluir projeto.", "Aberto"),
    ("David Valoski", "Book Papelaria", "Atenção", "Conclusão", "Conta resgatada voltou a usar, mas deixou de responder depois.", "David Valoski", "Esta semana", "Concluir por mensagem e compartilhar treinamentos gravados.", "Aberto"),
    ("David Valoski", "Distribuidora APJ", "Risco", "Projeto travado", "Projeto sem treinamento, com ghosting e boleto próximo do vencimento.", "David Valoski", "Urgente", "Puxar reunião de treinamento e validar bloqueio financeiro.", "Aberto"),
    ("David Valoski", "Sism", "Oportunidade", "Upgrade", "Cliente quer entender próximos planos após downgrade e precisa de passagem para expansão.", "David Valoski", "Esta semana", "Acionar time de expansão/inside sales.", "Aberto"),
    ("David Valoski", "Tempero Forte", "Atenção", "Projeto travado", "Conta só deve responder depois da implantação do ERP; depende de timing externo.", "David Valoski", "Monitoramento", "Adicionar Ana Cassia ao grupo Nomus e aguardar conclusão do ERP.", "Aberto"),

    ("Maria Elena", "Sagra", "Oportunidade", "Upgrade", "Discussão concentrou oportunidade de expansão via funil de vendas e convenção.", "Ana Cassia", "Esta semana", "Responder temas de expansão com Jorge e avaliar apoio na convenção.", "Aberto"),
    ("Maria Elena", "Distribuidora DPL", "Risco", "Projeto travado", "Prioridade máxima da semana, com go live cancelado e implantação Sankhya emperrada.", "Maria Elena", "Urgente", "Ligar para o cliente, aprofundar causa e escalar engenharia se necessário.", "Aberto"),
    ("Maria Elena", "Auris", "Risco", "Projeto travado", "Projeto antigo, em go live, pressionando por resposta de customização.", "Ana Cassia", "Até amanhã meio-dia", "Responder sobre orçamento de customização e definir próximo passo.", "Aberto"),
    ("Maria Elena", "Ouro Mineiro", "Risco", "Negociação de cancelamento", "Projeto emperrado por incompatibilidade estrutural com Questor e definição de CNPJs.", "Maria Elena", "Até amanhã", "Confirmar decisão do cliente sobre segundo CNPJ e viabilidade do projeto.", "Aberto"),
    ("Maria Elena", "RF", "Risco", "Negociação de cancelamento", "Conta em impasse operacional sobre recebimento de pedidos e com risco de não renovação.", "Maria Elena", "Esta semana", "Mapear com cliente e produto o que é possível desenvolver ou não.", "Aberto"),
    ("Maria Elena", "Trisquel Cosméticos", "Atenção", "Ressurrect", "Conta da base com histórico de troca de responsável e risco anterior, agora com caminho de retomada.", "Maria Elena", "Hoje", "Agendar treinamento de configurações e validar troca de integração.", "Aberto"),
    ("Maria Elena", "Doce Rico", "Atenção", "Conclusão", "Projeto em fase de encerramento dependendo de retorno do cliente após ajuste do parceiro.", "Maria Elena", "Esta semana", "Conversar com o cliente e conduzir encerramento formal.", "Aberto"),
    ("Maria Elena", "Gold Beauty", "Perdido", "Perdido", "Dois projetos sem integração devem ser encerrados formalmente.", "Maria Elena", "Esta semana", "Formalizar encerramento na operação.", "Aberto"),
    ("Maria Elena", "Cianite", "Saudável", "Conclusão", "Conta com resultado positivo e checkpoint de encerramento planejado.", "Maria Elena", "Esta semana", "Agendar checkpoint com Érica ou William e concluir.", "Aberto"),
    ("Maria Elena", "Top Frios", "Saudável", "Ressurrect", "Conta foi reativada na semana anterior e entra como vitória qualitativa da carteira.", "Maria Elena", "Monitoramento", "Acompanhar estabilidade pós-retomada.", "Aberto"),
    ("Maria Elena", "B Magalhães Agroflora", "Saudável", "Ressurrect", "Projeto reativado com bom sinal de recuperação da carteira.", "Maria Elena", "Monitoramento", "Ajustar cadastro no Salesforce e acompanhar retomada.", "Aberto"),

    ("Natieli Ambrosi", "Portal Cuiabá", "Perdido", "Perdido", "Cliente já cancelado; deve sair da agenda e ser finalizado como perdido.", "Natieli Ambrosi", "Imediato", "Retirar da agenda e registrar como perdido.", "Aberto"),
    ("Natieli Ambrosi", "TDP", "Risco", "Projeto travado", "Projeto em desenvolvimento com promessa de prazo, mas sem novos contatos do cliente.", "Natieli Ambrosi", "Até 18/05", "Cobrar Thiago e acompanhar entrega prometida por Pat.", "Aberto"),
    ("Natieli Ambrosi", "Injeção Eletrônica", "Risco", "Negociação de cancelamento", "Conta nunca usou Mercos, trocou equipe várias vezes e segue muito desorganizada.", "Ana Cassia", "Amanhã", "Tentar contato executivo com Carlos Eduardo para destravar ou confirmar risco.", "Aberto"),
    ("Natieli Ambrosi", "Saboretos", "Atenção", "Projeto travado", "Cliente segue empacado por configuração do novo ERP, mas sem cancelamento iminente declarado.", "Natieli Ambrosi", "Esta semana", "Checar status e manter acompanhamento sem pressionar a ponto de gerar churn.", "Aberto"),
    ("Natieli Ambrosi", "Reale", "Saudável", "Conclusão", "Troca de integração para CORP avançou bem e projeto caminha para encerramento.", "Natieli Ambrosi", "7-10 dias", "Remover CORP do grupo depois da reunião final e concluir.", "Aberto"),
    ("Natieli Ambrosi", "Suprva", "Risco", "Perdido", "Cliente só quer encerrar após funcionalidade específica, mas sem demanda ativa real.", "Natieli Ambrosi", "A definir", "Avaliar marcar como perdido se a inércia continuar.", "Aberto"),
    ("Natieli Ambrosi", "Mais Trama", "Risco", "Negociação de cancelamento", "Cliente inadimplente e sem retorno, com expectativa de cancelamento.", "Natieli Ambrosi", "Esta semana", "Seguir tratativa de cancelamento e atualizar status.", "Aberto"),
    ("Natieli Ambrosi", "ML Induzidos", "Saudável", "Conclusão", "Cliente finalmente usa o sistema e pode ser concluído apesar do silêncio recente.", "Natieli Ambrosi", "Esta semana", "Encerrar o projeto com apoio do contato do OMIE, se necessário.", "Aberto"),
    ("Natieli Ambrosi", "Atalaia", "Perdido", "Perdido", "Projeto orientado para ser marcado como perdido.", "Natieli Ambrosi", "Esta semana", "Registrar perda formalmente.", "Aberto"),
    ("Natieli Ambrosi", "Space", "Perdido", "Perdido", "Projeto orientado para ser marcado como perdido.", "Natieli Ambrosi", "Esta semana", "Registrar perda formalmente.", "Aberto"),
    ("Natieli Ambrosi", "Raiata 2", "Atenção", "Conclusão", "Conta secundária usa Mercos, mas não configurou B2B e não responde.", "Natieli Ambrosi", "Esta semana", "Concluir o projeto e alinhar pendência de hierarquia com Miri.", "Aberto"),
    ("Natieli Ambrosi", "Latomio", "Atenção", "Conclusão", "Conta secundária não interage; estratégia será vídeo + prazo curto antes de concluir.", "Natieli Ambrosi", "10 dias", "Enviar orientações, deixar rodar e concluir se não houver retorno.", "Aberto"),

    ("Aline Santos", "Big Sorvete", "Risco", "Negociação de cancelamento", "Provável cancelamento ligado à demora da integração com parceiro externo.", "Ana Cassia", "Esta semana", "Responder e-mail de finalização, envolver Fábio e registrar motivo com clareza.", "Aberto"),
    ("Aline Santos", "Bom Biscoito", "Oportunidade", "Upgrade", "Conta pequena saiu do radar da lista, mas reunião quer tentar recuperar MRR via expansão.", "Aline Santos", "Esta semana", "Mapear oportunidade de reversão e upgrade.", "Aberto"),
    ("Aline Santos", "Castro", "Risco", "Negociação de cancelamento", "Cliente grande e lento, com risco operacional por atraso e possível necessidade de estender ramp-up.", "Aline Santos", "Até 15 dias", "Negociar prazo final de implantação e usar mês extra de ramp-up apenas como último recurso.", "Aberto"),
    ("Aline Santos", "Ferraspar", "Atenção", "Sem movimento", "Não é risco de cancelamento direto, mas engajamento B2B está fraco e contato principal deu no-show.", "Aline Santos", "Esta semana", "Tentar outro gerente enquanto o responsável está de férias.", "Aberto"),
    ("Aline Santos", "Silvio Pires", "Atenção", "Projeto travado", "Conta dependia de entrega da Nomus; coordenação decidiu retirar do risco formal.", "Aline Santos", "Fim do mês", "Aguardar data prometida e revisar status depois da entrega.", "Aberto"),
    ("Aline Santos", "DCA", "Saudável", "Conclusão", "Cliente já usa o sistema e pode sair do risco para conclusão via WhatsApp.", "Aline Santos", "Esta semana", "Concluir projeto e pedir nota.", "Aberto"),
    ("Aline Santos", "Ricati", "Atenção", "Projeto travado", "Integração com Alterdata segue com exceção fiscal pendente fora do Mercos.", "Aline Santos", "Esta semana", "Pedir ao cliente que abra chamado com a Alterdata e compartilhar resposta.", "Aberto"),
    ("Aline Santos", "Tecom", "Perdido", "Perdido", "Projeto de baixo MRR, arrastado por oito meses e sem justificativa para seguir consumindo carteira.", "Aline Santos", "Esta semana", "Finalizar como perdido e comunicar possibilidade de retomada com novo cronograma.", "Aberto"),
    ("Aline Santos", "Tempero Forte", "Atenção", "Projeto travado", "Cliente não responde; operação quer apoio do David para engajar.", "David", "Esta semana", "Marcar reunião com o cliente e reaquecer contato.", "Aberto"),
    ("Aline Santos", "Duas Rodas Motopeça", "Atenção", "Expansão", "Cliente mudou para ERP não parceiro; há alternativa de uso sem integração ou com integradora BLP.", "Aline Santos", "Esta semana", "Apresentar caminhos comerciais e operacionais para manter valor.", "Aberto"),
    ("Aline Santos", "Night Toy", "Risco", "Perdido", "Conta recebeu prazo final curto e deve ser encerrada se continuar sem retorno.", "Aline Santos", "1 semana", "Dar prazo máximo e encerrar se o cliente não responder.", "Aberto"),
]


def parse_docx_paragraphs(path: Path) -> list[str]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    paragraphs: list[str] = []
    for node in root.findall(".//w:p", ns):
        text = "".join(part.text or "" for part in node.findall(".//w:t", ns)).strip()
        if text:
            paragraphs.append(text)
    return paragraphs


def extract_next_steps(paragraphs: list[str]) -> list[str]:
    lines: list[str] = []
    in_steps = False
    for paragraph in paragraphs:
        if paragraph == "Próximas etapas":
            in_steps = True
            continue
        if not in_steps:
            continue
        if paragraph == "Detalhes":
            break
        if paragraph.startswith("["):
            lines.append(paragraph)
    return lines


def infer_client_name(text: str) -> str:
    normalized = text
    patterns = [
        r"cliente ([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ][^:.;]+)",
        r"projeto da ([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ][^:.;]+)",
        r"projeto ([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ][^:.;]+)",
        r"conta do cliente ([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ][^:.;]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE)
        if match:
            return clean_label(match.group(1))
    return "Tema geral"


def clean_label(value: str) -> str:
    return (
        value.replace("*", "")
        .replace("  ", " ")
        .strip(" .;:-")
        .strip()
    )


def parse_action_line(line: str, implanter: str, meeting_date: str, source_file: str) -> dict[str, str]:
    match = re.match(r"^\[(?P<owner>[^\]]+)\]\s*(?P<action>[^:]+):\s*(?P<description>.+)$", line)
    if not match:
        return {
            "semana": WEEK_LABEL,
            "data_batida": meeting_date,
            "implanter": implanter,
            "responsavel": implanter,
            "acao": "Sem parsing",
            "cliente_tema": "Tema geral",
            "descricao": line,
            "prazo": "",
            "origem": source_file,
        }

    description = clean_label(match.group("description"))
    deadline = infer_deadline(description)
    return {
        "semana": WEEK_LABEL,
        "data_batida": meeting_date,
        "implanter": implanter,
        "responsavel": clean_label(match.group("owner")),
        "acao": clean_label(match.group("action")),
        "cliente_tema": infer_client_name(description),
        "descricao": description,
        "prazo": deadline,
        "origem": source_file,
    }


def infer_deadline(text: str) -> str:
    lowered = text.lower()
    if "hoje" in lowered:
        return "Hoje"
    if "amanhã" in lowered or "amanha" in lowered:
        return "Amanhã"
    if "quarta" in lowered:
        return "Quarta-feira"
    if "semana" in lowered:
        return "Esta semana"
    if "quinta" in lowered:
        return "Quinta-feira"
    return ""


def write_section_title(ws, row: int, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="left")
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=9)
        cell = ws.cell(row=row + 1, column=1, value=subtitle)
        cell.font = Font(size=11, color="4B5563")
        cell.fill = PatternFill("solid", fgColor="EAF2FF")
    return row + (2 if subtitle else 1)


def apply_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    table = Table(displayName=name, ref=f"A{start_row}:{get_column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_headers(ws, row: int, columns: Iterable[str]) -> None:
    thin = Side(style="thin", color="D1D5DB")
    for index, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def autofit(ws) -> None:
    widths: dict[int, int] = defaultdict(lambda: 12)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = min(max(widths[cell.column], len(value) + 2), 45)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> None:
    paragraphs_by_doc = {doc["path"].name: parse_docx_paragraphs(doc["path"]) for doc in DOCS}
    action_rows: list[dict[str, str]] = []
    for doc in DOCS:
        source_name = doc["path"].name
        steps = extract_next_steps(paragraphs_by_doc[source_name])
        for line in steps:
            action_rows.append(parse_action_line(line, doc["implanter"], doc["meeting_date"], source_name))

    qualitative_rows = []
    for implanter, client, criticality, movement, summary, owner, deadline, next_step, action_status in QUALITATIVE_ROWS:
        qualitative_rows.append(
            {
                "semana": WEEK_LABEL,
                "data_batida": "2026-05-11",
                "implanter": implanter,
                "cliente_tema": client,
                "tipo_registro": "Cliente",
                "criticidade_sugerida": criticality,
                "movimento_forecast": movement,
                "situacao_resumida": summary,
                "proximo_passo": next_step,
                "responsavel": owner,
                "prazo": deadline,
                "observacao": "Extraído das anotações do Gemini e estruturado para a planilha qualitativa.",
                "mrr_impactado": None,
                "status_acao": action_status,
            }
        )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo Semanal"
    ws_qual = wb.create_sheet("Planilha Qualitativa")
    ws_actions = wb.create_sheet("Ações da Semana")
    ws_dict = wb.create_sheet("Dicionário")

    row = 1
    row = write_section_title(
        ws_summary,
        row,
        "Batidas de funil consolidadas da semana",
        "Versão 1 da 8ª planilha qualitativa montada a partir das 5 batidas recebidas. A batida da Samara ainda está pendente.",
    )
    row += 1

    summary_cards = [
        ("Batidas recebidas", len(DOCS)),
        ("Implanters cobertos", len({doc["implanter"] for doc in DOCS})),
        ("Registros qualitativos", len(qualitative_rows)),
        ("Ações extraídas", len(action_rows)),
        ("Batida pendente", "Samara"),
    ]
    for idx, (label, value) in enumerate(summary_cards, start=1):
        c1 = ws_summary.cell(row=row, column=idx * 2 - 1, value=label)
        c2 = ws_summary.cell(row=row + 1, column=idx * 2 - 1, value=value)
        c1.font = Font(bold=True, color="1E3A5F")
        c1.fill = PatternFill("solid", fgColor="DBEAFE")
        c2.font = Font(size=15, bold=True, color="111827")
        c2.fill = PatternFill("solid", fgColor="EFF6FF")
        ws_summary.merge_cells(start_row=row, start_column=idx * 2 - 1, end_row=row, end_column=idx * 2)
        ws_summary.merge_cells(start_row=row + 1, start_column=idx * 2 - 1, end_row=row + 1, end_column=idx * 2)
    row += 4

    movement_counts = Counter(item["movimento_forecast"] for item in qualitative_rows)
    criticality_counts = Counter(item["criticidade_sugerida"] for item in qualitative_rows)
    by_implanter = Counter(item["implanter"] for item in qualitative_rows)

    ws_summary.cell(row=row, column=1, value="Movimento forecast")
    ws_summary.cell(row=row, column=3, value="Qtd.")
    ws_summary.cell(row=row, column=5, value="Criticidade sugerida")
    ws_summary.cell(row=row, column=7, value="Qtd.")
    for cell_ref in ("A", "C", "E", "G"):
        ws_summary[f"{cell_ref}{row}"].font = Font(bold=True, color="FFFFFF")
        ws_summary[f"{cell_ref}{row}"].fill = PatternFill("solid", fgColor="0F766E")
    row += 1
    max_rows = max(len(movement_counts), len(criticality_counts))
    movement_items = sorted(movement_counts.items(), key=lambda item: (-item[1], item[0]))
    criticality_items = sorted(criticality_counts.items(), key=lambda item: (-item[1], item[0]))
    for offset in range(max_rows):
        if offset < len(movement_items):
            label, count = movement_items[offset]
            ws_summary.cell(row=row + offset, column=1, value=label)
            ws_summary.cell(row=row + offset, column=3, value=count)
        if offset < len(criticality_items):
            label, count = criticality_items[offset]
            ws_summary.cell(row=row + offset, column=5, value=label)
            ws_summary.cell(row=row + offset, column=7, value=count)
    row += max_rows + 2

    ws_summary.cell(row=row, column=1, value="Implanter")
    ws_summary.cell(row=row, column=2, value="Registros")
    ws_summary.cell(row=row, column=3, value="Observação")
    for col in range(1, 4):
        ws_summary.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws_summary.cell(row=row, column=col).fill = PatternFill("solid", fgColor="7C3AED")
    row += 1
    for implanter, count in sorted(by_implanter.items()):
        ws_summary.cell(row=row, column=1, value=implanter)
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(
            row=row,
            column=3,
            value="Samara ainda não enviada" if implanter == "Samara" else "Batida recebida e consolidada",
        )
        row += 1
    ws_summary.cell(row=row, column=1, value="Samara")
    ws_summary.cell(row=row, column=2, value=0)
    ws_summary.cell(row=row, column=3, value="Faltando receber a batida da semana")

    qual_headers = [
        "Semana",
        "Data da batida",
        "Implanter",
        "Cliente / tema",
        "Tipo de registro",
        "Criticidade sugerida",
        "Movimento forecast",
        "Situação resumida",
        "Próximo passo",
        "Responsável",
        "Prazo",
        "Observação",
        "MRR impactado (preencher)",
        "Status da ação",
    ]
    style_headers(ws_qual, 1, qual_headers)
    for row_idx, item in enumerate(qualitative_rows, start=2):
        values = [
            item["semana"],
            item["data_batida"],
            item["implanter"],
            item["cliente_tema"],
            item["tipo_registro"],
            item["criticidade_sugerida"],
            item["movimento_forecast"],
            item["situacao_resumida"],
            item["proximo_passo"],
            item["responsavel"],
            item["prazo"],
            item["observacao"],
            item["mrr_impactado"],
            item["status_acao"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_qual.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_table(ws_qual, 1, len(qualitative_rows) + 1, len(qual_headers), "QualitativeWeeklyTable")
    ws_qual.freeze_panes = "A2"

    action_headers = [
        "Semana",
        "Data da batida",
        "Implanter",
        "Responsável",
        "Ação",
        "Cliente / tema",
        "Descrição",
        "Prazo",
        "Origem",
    ]
    style_headers(ws_actions, 1, action_headers)
    for row_idx, item in enumerate(action_rows, start=2):
        values = [
            item["semana"],
            item["data_batida"],
            item["implanter"],
            item["responsavel"],
            item["acao"],
            item["cliente_tema"],
            item["descricao"],
            item["prazo"],
            item["origem"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_actions.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_table(ws_actions, 1, len(action_rows) + 1, len(action_headers), "WeeklyActionsTable")
    ws_actions.freeze_panes = "A2"

    dict_headers = ["Campo", "Uso recomendado", "Opções sugeridas"]
    style_headers(ws_dict, 1, dict_headers)
    dictionary_rows = [
        ("Criticidade sugerida", "Leitura qualitativa decidida na batida", "Risco | Atenção | Saudável | Perdido | Oportunidade"),
        ("Movimento forecast", "Evento que deve alimentar a futura 4ª página de forecast", "Negociação de cancelamento | Projeto travado | Perdido | Revertido | Ressurrect | Upgrade | Contraction | Conclusão | Expansão | Sem movimento"),
        ("MRR impactado (preencher)", "Campo manual para enriquecer forecast", "Valor em R$ sem separador de milhar"),
        ("Status da ação", "Controle semanal do follow-up", "Aberto | Em andamento | Concluído | Aguardando cliente"),
        ("Observação", "Memória operacional da batida", "Texto livre, idealmente curto e objetivo"),
    ]
    for row_idx, values in enumerate(dictionary_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws_dict.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_table(ws_dict, 1, len(dictionary_rows) + 1, len(dict_headers), "QualitativeDictionary")

    dv_criticality = DataValidation(
        type="list",
        formula1='"Risco,Atenção,Saudável,Perdido,Oportunidade"',
        allow_blank=True,
    )
    dv_movement = DataValidation(
        type="list",
        formula1='"Negociação de cancelamento,Projeto travado,Perdido,Revertido,Ressurrect,Upgrade,Contraction,Conclusão,Expansão,Sem movimento"',
        allow_blank=True,
    )
    dv_action = DataValidation(
        type="list",
        formula1='"Aberto,Em andamento,Concluído,Aguardando cliente"',
        allow_blank=True,
    )
    ws_qual.add_data_validation(dv_criticality)
    ws_qual.add_data_validation(dv_movement)
    ws_qual.add_data_validation(dv_action)
    dv_criticality.add(f"F2:F{len(qualitative_rows) + 1}")
    dv_movement.add(f"G2:G{len(qualitative_rows) + 1}")
    dv_action.add(f"N2:N{len(qualitative_rows) + 1}")

    ws_qual.conditional_formatting.add(
        f"F2:F{len(qualitative_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Risco"'], fill=PatternFill("solid", fgColor="FEE2E2")),
    )
    ws_qual.conditional_formatting.add(
        f"F2:F{len(qualitative_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Atenção"'], fill=PatternFill("solid", fgColor="FEF3C7")),
    )
    ws_qual.conditional_formatting.add(
        f"F2:F{len(qualitative_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Saudável"'], fill=PatternFill("solid", fgColor="DCFCE7")),
    )

    for ws in (ws_summary, ws_qual, ws_actions, ws_dict):
        autofit(ws)

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 14
    ws_summary.column_dimensions["E"].width = 24
    ws_summary.column_dimensions["G"].width = 14

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Saved workbook to {OUTPUT}")


if __name__ == "__main__":
    main()
